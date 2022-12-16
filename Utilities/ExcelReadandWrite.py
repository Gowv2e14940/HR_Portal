import openpyxl
import xlrd
import pandas
def getRowCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.max_row

def getColumnCount(file, sheetName):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.max_column

def readData(file, sheetName, rownum, columnno):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        return sheet.cell(row=rownum, column=columnno).value

def writeData(file, sheetName, rownum, columnno, data):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.get_sheet_by_name(sheetName)
        sheet.cell(row=rownum, column=columnno).value = data
        workbook.save(file)


def excelsearch(file, sheetname , data):

    # book = xlrd.open_workbook(file)
    book = openpyxl.load_workbook(file)
    sheet = book.get_sheet_by_name(sheetname)

    for rowidx in range(1,sheet.max_row+1):
        # row = sheet.row(rowidx)
        # for colidx, cell in enumerate(sheet.row(rowidx)):
        for colidx in range(1,sheet.max_column+1):
            if str(sheet.cell(rowidx,colidx).value).__contains__(data):
                print(rowidx)
                print("The data is", str(sheet.cell(rowidx,colidx).value))
                return rowidx